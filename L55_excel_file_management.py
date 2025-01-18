import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('Student.xlsx')
sheet = wb['Session22_25']
# cell = sheet['a1']
#       or
cell = sheet.cell(1, 1)
print(cell.value)
no_of_rows = sheet.max_row
print(no_of_rows)
for i in range(1, no_of_rows+1):
    cell = sheet.cell(i, 2)
    print(cell.value)
for i in range(2, no_of_rows+1):
    updated_value = ''
    updated_value_cell = sheet.cell(i, 6)
    updated_value_cell.value = updated_value
wb.save('Student.xlsx')

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   max_col=5,
                   min_col=5)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'j2')
wb.save('Student.xlsx')
